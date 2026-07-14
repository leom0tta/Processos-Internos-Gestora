import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# pip install openpyxl requests

API_KEY = os.getenv('GORILA_API_KEY')
BASE_URL = os.getenv('GORILA_API_BASE_URL')

start = "2026-06-01"
end = "2026-06-30"

# ── Feriados do período ───────────────────────────────────────
feriados = [
    "2026-06-04",  # Corpus Christi
]
# ─────────────────────────────────────────────────────────────

headers_api = {"authorization": API_KEY}

def listar_todos_portfolios():
    """Busca todos os portfólios da conta com paginação automática."""
    portfolios = []
    page_token = None

    while True:
        params = {"limit": 1000}
        if page_token:
            params["pageToken"] = page_token

        resp = requests.get(f"{BASE_URL}/portfolios", headers=headers_api, params=params, timeout=30)
        resp.raise_for_status()
        data = resp.json()

        portfolios.extend(data.get("records", []))
        print(f"  Portfolios carregados: {len(portfolios)}")

        # next é None ou string vazia quando não há mais páginas
        next_url = data.get("next")
        if not next_url:
            break

        # Extrai o pageToken da URL de próxima página
        from urllib.parse import urlparse, parse_qs
        token = parse_qs(urlparse(next_url).query).get("token", [None])[0]
        if not token:
            break
        page_token = token

    return portfolios

def is_dia_util(date_str):
    d = datetime.strptime(date_str, "%Y-%m-%d").date()
    return d.weekday() < 5 and date_str not in feriados

def get_nav_diario(portfolio_id, start=start, end=end):
    url = f"{BASE_URL}/portfolios/{portfolio_id}/nav"
    params = {"frequency": "DAILY", "startDate": start, "endDate": end}
    resp = requests.get(url, headers=headers_api, params=params, timeout=30)
    resp.raise_for_status()
    return resp.json()

# ── Busca portfolios automaticamente ─────────────────────────
print("Listando portfólios da conta Gorila...")
portfolios = listar_todos_portfolios()
print(f"Total encontrado: {len(portfolios)} portfólios\n")

# ── Coleta NAV de cada portfolio ──────────────────────────────
linhas = []
for p in portfolios:
    nome = p.get("name", p["id"])
    pid  = p["id"]
    status = p.get("status", "")

    # Ignora portfolios com erro de processamento
    if status == "INVALID":
        print(f"  ⚠️  Ignorando {nome} (status: INVALID)")
        continue

    print(f"Buscando: {nome}...")
    try:
        data = get_nav_diario(pid)
        serie = data.get("timeseries", [])

        uteis = [
            pt for pt in serie
            if pt["nav"] is not None and is_dia_util(pt["referenceDate"])
        ]
        navs = [pt["nav"] for pt in uteis]

        linhas.append({
            "Cliente":               nome,
            "ID":                    pid,
            "Dias úteis com dados":  len(navs),
            "Mínimo (R$)":           min(navs) if navs else None,
            "Máximo (R$)":           max(navs) if navs else None,
            "Média pro rata (R$)":   sum(navs) / len(navs) if navs else None,
            "Último dia útil":       uteis[-1]["referenceDate"] if uteis else None,
            "Patrimônio final (R$)": uteis[-1]["nav"] if uteis else None,
        })
    except Exception as e:
        print(f"  ⚠️  Erro em {nome}: {e}")
        linhas.append({
            "Cliente": nome, "ID": pid,
            "Dias úteis com dados": None, "Mínimo (R$)": None, "Máximo (R$)": None,
            "Média pro rata (R$)": None, "Último dia útil": None, "Patrimônio final (R$)": None,
        })

# ── Gera Excel ────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Patrimônio junho 2026"

colunas = ["Cliente", "ID", "Dias úteis com dados", "Mínimo (R$)",
           "Máximo (R$)", "Média pro rata (R$)", "Último dia útil", "Patrimônio final (R$)"]

header_fill = PatternFill("solid", fgColor="1E3A5F")
header_font = Font(bold=True, color="FFFFFF", size=11)
for col_idx, col_name in enumerate(colunas, 1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

moeda_fmt = 'R$ #,##0.00'
for row_idx, linha in enumerate(linhas, 2):
    ws.cell(row=row_idx, column=1, value=linha["Cliente"])
    ws.cell(row=row_idx, column=2, value=linha["ID"]).font = Font(size=9, color="888888")
    ws.cell(row=row_idx, column=3, value=linha["Dias úteis com dados"]).alignment = Alignment(horizontal="center")
    for col, key in [(4, "Mínimo (R$)"), (5, "Máximo (R$)"),
                     (6, "Média pro rata (R$)"), (8, "Patrimônio final (R$)")]:
        cell = ws.cell(row=row_idx, column=col, value=linha[key])
        cell.number_format = moeda_fmt
    ws.cell(row=row_idx, column=7, value=linha["Último dia útil"]).alignment = Alignment(horizontal="center")
    if row_idx % 2 == 0:
        for col in range(1, len(colunas) + 1):
            ws.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor="F0F4F8")

larguras = [40, 38, 20, 18, 18, 22, 16, 22]
for i, w in enumerate(larguras, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 22
ws.freeze_panes = "A2"

wb.save("patrimonio_pro_rata_junho_2026.xlsx")
print(f"\n✅ {len(linhas)} portfólios exportados → patrimonio_pro_rata_junho_2026.xlsx")